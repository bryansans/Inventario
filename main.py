import speech_recognition as sr
import openpyxl
from reportlab.pdfgen import canvas
import matplotlib.pyplot as plt
from tkinter import Tk, Label, Entry, Button, Text, END, StringVar
from datetime import datetime

def reconocer_voz():
    # TODO: Implementar la lógica para reconocer voz
    recognizer = sr.Recognizer()
    mic = sr.Microphone()
    
    with mic as source:
    print("Habla ahora...")
    Recognizer.adjust_for_ambient_noise(source)
    audio = Recognizer.listen(source)

    try:
    print("Reconociendo...")
    texto = recognizer.recognize_google(audio, language='es-ES')
    print(f"Texto reconocido: {texto}")
    return texto
    except sr.UnknownValueError:
    print("No se pudo entender el audio")
    return ""
    except sr.RequestError:
    print("Error en la solicitud")
    return ""


def exportar_a_excel(inventario):
    # TODO: Implementar la lógica para exportar a Excel
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Inventario"

    sheet['A1'] = 'Producto'
    sheet['B1'] = 'Cantidad'

    row = 2
    for producto, cantidad in inventario ¿.items():
    sheet.cell(row=row, column=1).value = producto
    sheet.cell(row=row, column=2).value = cantidad
    row += 1

    wb.save('inventario.xlsx')
    print("Inventario exportado a Excel exitosamente.")


def exportar_a_pdf(inventario):
    # TODO: Implementar la lógica para exportar a PDF
    pdf = canvas.Canvas("inventario.pdf")
    pdf.drawString(100, 800, "Inventario")

    y = 750
    for producto, cantidad in inventario.items():
    pdf.drawString(100, y, f"Producto: {producto}, Cantidad: {cantidad}")
    y -= 20

    pdf.save()
    print("Inventario exportado a PDF exitosamente.")


def mostrar_grafico_salidas(inventario):
    # TODO: Implementar la lógica para mostrar gráfico de salidas
    productos = list(inventario.keys())
    cantidades = list(inventario.values())

    plt.bar(productos, cantidades)
    plt.xlabel('Productos')
    plt.ylabel('Cantidad')
    plt.title('Inventario de Productos')
    plt.xticks(rotation=45)
    plt.tight_layout()

    plt.show()


def main():
    inventario = {}

    root = Tk()
    root.title("Registro de Inventario")

    # Campos de entrada
    producto_var = StringVar()
    tipo_var = StringVar()

    Label(root, text="Producto:").grid(row=0, column=0)
    Entry(root, textvariable=producto_var).grid(row=0, column=1)


    # Botones
    Button(root, text="Registrar desde GUI", command=lambda: registrar_desde_gui(producto_var, cantidad_var, "añadir el resto")).grid(row=4, column=0, columnspan=2)


    # Widget de texto para mostrar el inventario
    text_widget = Text(root, height=10, width=50)
    text_widget.grid(row=9, column=0, columnspan=2)

    # Mostrar el inventario inicialmente
    mostrar_inventario(inventario, text_widget)

    root.mainloop()

if __name__ == "__main__":
    main()
